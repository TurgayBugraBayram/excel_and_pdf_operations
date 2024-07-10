import pandas as pd
from openpyxl import load_workbook

file_path = 'PA-Forms_Tool.xlsx'
update_file_path = 'update.xlsx'
df = pd.read_excel(file_path, header=1)

wb = load_workbook(file_path)
sheet = wb.active

columns = ['AK', 'IA', 'ID', 'IL', 'MI', 'MN', 'MT', 'ND', 'NV', 'OR', 'SD', 'UT', 'WA', 'WI']

selected_state = input("Please enter a state name (e.g., 'AK'): ")

if selected_state not in columns:
    print("Invalid state name. Please enter a valid state name.")
else:
    col_idx = df.columns.get_loc(selected_state) + 1

    count_X_delete = 0
    count_X = 0
    form_data_X_delete = []
    form_data_X = []

    for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
        cell = row[0]
        cell_value = cell.value

        if cell_value in ['X', 'x']:
            if cell.font and cell.font.strike:
                if cell.row - 2 < len(df):
                    count_X_delete += 1
                    form_data_X_delete.append((cell.row, df.loc[cell.row - 2, ['Form number w/Edition date', 'Form Name']]))
            else:
                if cell.row - 2 < len(df):
                    count_X += 1
                    form_data_X.append((cell.row, df.loc[cell.row - 2, ['Form number w/Edition date', 'Form Name','GW Document Type']]))


    print(f"\nThere are {count_X_delete} crossed out 'X' or 'x' and {count_X} non-crossed out 'X' or 'x' in the {selected_state} column.")

    if count_X_delete > 0:
        print("\nForms with crossed out 'X' or 'x':")
        for row_num, data in form_data_X_delete:
            form_number, form_name = data['Form number w/Edition date'], data['Form Name']
            print(f"{row_num}: Form number w/Edition date: {form_number}, Form Name: {form_name}")
    else:
        print(f"No crossed out 'X' or 'x' found in the {selected_state} column.")

    if count_X > 0:
        print("\nForms with non-crossed out 'X' or 'x':")
        for row_num, data in form_data_X:
            form_number, form_name, form_GW = data['Form number w/Edition date'], data['Form Name'], data['GW Document Type']
            print(f"{row_num}: Form number w/Edition date: {form_number}, Form Name: {form_name} {form_GW}")
    else:
        print(f"No non-crossed out 'X' or 'x' found in the {selected_state} column.")

df_filtered =pd.DataFrame([data[1] for data in form_data_X])

with pd.ExcelWriter(update_file_path,engine="openpyxl") as writer:
    df_filtered.to_excel(writer, index=False)