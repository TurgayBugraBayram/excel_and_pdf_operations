import pandas as pd
from openpyxl import load_workbook

file_path = 'PA-Forms_Tool.xlsx'
df = pd.read_excel(file_path, header=1)

wb = load_workbook(file_path)
sheet = wb.active

columns = ['AK', 'IA', 'ID', 'IL', 'MI', 'MN', 'MT', 'ND', 'NV', 'OR', 'SD', 'UT', 'WA', 'WI']

selected_state = input("Lütfen bir eyalet adı girin (örneğin 'AK'): ")

if selected_state not in columns:
    print("Geçersiz eyalet adı. Lütfen doğru bir eyalet adı girin.")
else:
    col_idx = df.columns.get_loc(selected_state) + 1

    count_X_delete = 0
    form_data_X_delete = []

    for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
        cell = row[0]
        cell_value = cell.value
        if cell_value in ['X', 'x']:  
            if cell.font and cell.font.strike:
                count_X_delete += 1
                form_data_X_delete.append((cell.row, df.loc[cell.row - 2, ['Form number w/Edition date', 'Form Name']]))

    count_X = df[selected_state].value_counts().get('X', 0) + df[selected_state].value_counts().get('x', 0)

    print(f"\n{selected_state} sütununda {count_X_delete} tane üstü çizili 'X' veya 'x' var ve {count_X - count_X_delete} 'X' veya 'x' var.")

    if count_X_delete > 0:
        print("\nÜstü çizili 'X' veya 'x' içeren formlar:")
        for row_num, data in form_data_X_delete:
            form_number, form_name = data['Form number w/Edition date'], data['Form Name']
            print(f"{row_num}: Form number w/Edition date: {form_number}, Form Name: {form_name}")
    else:
        print(f"{selected_state} sütununda üstü çizili 'X' veya 'x' bulunamadı.")
